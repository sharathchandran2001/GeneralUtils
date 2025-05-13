import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

st.title("Excel Sheet Search and Viewer")

# Upload Excel file
uploaded_file = st.file_uploader("Upload an Excel file (.xlsx)", type=["xlsx"])

def read_all_cells(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = [str(cell).strip() if cell is not None else "" for cell in row]
        data.append(row_data)
    return data

def flatten(data):
    return [cell for row in data for cell in row if cell]

# Input text for search
search_text = st.text_input("Enter text to search in Sheet1 to Sheet4")

# Add buttons
search_button = st.button("Search")
display_button = st.button("Display All Contents")

# Main logic
if uploaded_file:
    wb = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)

    if display_button:
        st.subheader("Contents of Sheet1 to Sheet4")
        for sheet_name in ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                values = read_all_cells(sheet)
                st.markdown(f"### {sheet_name}")
                st.dataframe(values)
            else:
                st.warning(f"{sheet_name} not found.")

    if search_button and search_text:
        st.subheader(f"Searching for '{search_text}' in Sheet1 to Sheet4...")
        found_in_sheets = []
        for sheet_name in ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                flat_data = flatten(read_all_cells(sheet))
                if any(search_text.lower() in str(cell).lower() for cell in flat_data):
                    found_in_sheets.append(sheet_name)

        if found_in_sheets:
            st.success(f"'{search_text}' found in: {', '.join(found_in_sheets)}")
        else:
            st.error(f"'{search_text}' not found in any of the sheets.")
