import sys
from openpyxl import load_workbook

##python rpa_xlsx_data_analyst.py Test.xlsx "Sheet"




def read_all_cells(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = [str(cell).strip() if cell is not None else "" for cell in row]
        data.append(row_data)
    return data

def flatten(data):
    return [cell for row in data for cell in row if cell]

def search_excel(file_path, search_text):
    try:
        wb = load_workbook(filename=file_path, data_only=True)
    except Exception as e:
        print(f"Failed to load Excel file: {e}")
        return

    found_in_sheets = []

    for sheet_name in ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            flat_data = flatten(read_all_cells(sheet))
            if any(search_text.lower() in cell.lower() for cell in flat_data):
                found_in_sheets.append(sheet_name)
        else:
            print(f"{sheet_name} not found in workbook.")

    if found_in_sheets:
        print(f"'{search_text}' found in: {', '.join(found_in_sheets)}")
    else:
        print(f"'{search_text}' not found in any of the sheets.")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python search_excel.py <path_to_excel_file.xlsx> <search_text>")
    else:
        excel_path = sys.argv[1]
        search_text = sys.argv[2]
        search_excel(excel_path, search_text)
